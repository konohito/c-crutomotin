#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""InBody(LookinBody Excel)データを台帳の測定に突合して Firestore へ取り込む ETL。

InBody の書き出しには氏名・参加者IDが無いため、以下のキーで記録用紙側の測定に突合する:
    測定年(テスト日の年) + 性別 + 身長(±HEIGHT_TOL) + 体重(±WEIGHT_TOL)
複数候補があるときは 地区(行政区) → 年齢 → 身長体重の近さ の順で一意化する。

突合できた測定ドキュメント measurements/* に `inbody` を merge 保存する:
    inbody = { smm, smi, fatPct, score, height, weight, ibId, district, testDate, source }

ファイルは Google Drive フォルダから自動取得する（ローカルに置く手間なし）。
Drive も Firestore も、利用者のローカルの ADC 認証で読み書きする。

  # 1) ADC に Drive 読み取りスコープを付けて認証（1回だけ）
  gcloud auth application-default login \
      --scopes=https://www.googleapis.com/auth/cloud-platform,https://www.googleapis.com/auth/drive.readonly

  # 2) 依存パッケージ
  pip install google-api-python-client google-auth google-cloud-firestore openpyxl

  # 3) ドライラン（Drive から取得 → 突合、書き込みなし）
  python etl-inbody.py --project cruto-motion

  # 4) 問題なければ書き込み
  python etl-inbody.py --project cruto-motion --commit

Drive を使わずローカルの xlsx で回したい場合は  --dir ./inbody  を指定する。
ドライランは Firestore を一切書き換えない。まず突合率を確認してから --commit すること。
"""
import argparse
import csv
import glob
import io
import os
import re
import sys
import tempfile

# InBody データが入っている Drive フォルダ
DEFAULT_FOLDER_ID = '111fACqvpnuz3mn2Et10ru28zTFf5NB8R'
SCOPES = ['https://www.googleapis.com/auth/cloud-platform',
          'https://www.googleapis.com/auth/drive.readonly']

HEIGHT_TOL = 2.0   # cm
WEIGHT_TOL = 1.5   # kg

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
# テスト用の見本データ行（機器の動作確認）は取り込まない
SKIP_IDS = {'cruto1', 'otamesi', '0000000888', '00001'}
SKIP_NAMES = {'otamesi', 'test', 'テスト', 'サンプル'}


def _norm_header(cell):
    return re.sub(r'^\s*\d+\.\s*', '', str(cell or '')).strip()


def find_cols(header):
    """ヘッダー行(番号付き)から、列名キーワードで各項目の列位置を特定する。
       LookinBody の書式ごとに列番号が違う（旧: 6.体重/27.骨格筋量、新: 15.体重/36.骨格筋量）ため
       番号ではなく名前で拾う。"""
    cols = {}
    for ci, cell in enumerate(header):
        h = _norm_header(cell)
        hl = h.lower()
        def put(k):
            cols.setdefault(k, ci)
        if hl == 'name' or h in ('氏名', '名前'):
            put('name')
        elif hl == 'id':
            put('id')
        elif ('height' in hl or '身長' in h) and 'limit' not in hl and 'range' not in hl and '標準' not in h:
            put('height')
        elif 'gender' in hl or '性別' in h:
            put('gender')
        elif hl == 'age' or h == '年齢':
            put('age')
        elif 'test date' in hl or '測定日' in h:
            put('testDate')
        elif hl == 'weight' or h == '体重':
            put('weight')
        elif ('skeletal muscle mass' in hl or '骨格筋量' in h) and 'limit' not in hl and 'range' not in hl and '/wt' not in hl:
            put('smm')
        elif 'skeletal muscle index' in hl or '骨格筋指数' in h or hl == 'smi':
            put('smi')
        elif ('percent body fat' in hl or '体脂肪率' in h or hl == 'pbf') and 'limit' not in hl and 'range' not in hl:
            put('fatPct')
        elif 'inbody score' in hl or 'inbody 点数' in h or h == '点数':
            put('score')
    return cols


def to_float(v):
    if v is None:
        return None
    try:
        s = str(v).strip().replace(',', '')
        if s in ('', '-'):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_year_date(raw):
    """'2023.09.11 15:06:10' -> (2023, '2023/09/11')"""
    m = re.search(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', str(raw or ''))
    if not m:
        return None, None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return y, f'{y}/{mo:02d}/{d:02d}'


# ---- Drive 取得 -------------------------------------------------------------
def fetch_from_drive(folder_id, creds):
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    drive = build('drive', 'v3', credentials=creds, cache_discovery=False)
    res = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields='files(id,name,mimeType,shortcutDetails)', pageSize=1000).execute()
    tmp = tempfile.mkdtemp(prefix='inbody_')
    paths = []
    for f in res.get('files', []):
        fid, name, mime = f['id'], f['name'], f['mimeType']
        if mime == 'application/vnd.google-apps.shortcut':
            sd = f.get('shortcutDetails') or {}
            fid = sd.get('targetId') or fid
            mime = sd.get('targetMimeType') or XLSX_MIME
        if 'spreadsheet' not in mime and not name.lower().endswith('.xlsx'):
            continue
        if not name.lower().endswith('.xlsx'):
            name += '.xlsx'
        req = (drive.files().export_media(fileId=fid, mimeType=XLSX_MIME)
               if mime == 'application/vnd.google-apps.spreadsheet'
               else drive.files().get_media(fileId=fid))
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        path = os.path.join(tmp, re.sub(r'[\\/]', '_', name))
        with open(path, 'wb') as fp:
            fp.write(buf.getvalue())
        paths.append(path)
    return paths


# ---- xlsx / csv 解析 --------------------------------------------------------
def _has_numbered_header(rows):
    for row in rows[:10]:
        if sum(1 for c in row if re.match(r'^\s*\d+\.', str(c or ''))) >= 8:
            return True
    return False


def read_rows(path):
    """xlsx でも csv でも行(list of list)を返す。xlsx は番号ヘッダーのあるシートを選ぶ。"""
    if path.lower().endswith('.csv'):
        import csv as _csv
        with open(path, newline='', encoding='utf-8-sig') as fp:
            return [list(r) for r in _csv.reader(fp)]
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    best = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if _has_numbered_header(rows):
            wb.close()
            return rows
        if len(rows) > len(best):
            best = rows
    wb.close()
    return best


def parse_rows(rows):
    """番号付きヘッダー行を探し、列名キーワードで列位置を特定してデータ行を取り出す。
       旧書式(令和4/5: 地区+機械ID)と新書式(2024/2025: 氏名+台帳ID)の両方に対応。"""
    header_i, cols = -1, {}
    for i, row in enumerate(rows[:10]):
        n_numbered = sum(1 for c in row if re.match(r'^\s*\d+\.', str(c or '')))
        if n_numbered >= 8:
            c = find_cols(row)
            # 必須列が揃うヘッダーのみ採用（血圧など別表のヘッダーを除外）
            if all(k in c for k in ('height', 'weight', 'testDate')):
                header_i, cols = i, c
                break
    if header_i < 0:
        return []

    # 旧書式は "1. ID" の左隣の無名列が地区(行政区)
    district_col = None
    if 'name' not in cols and cols.get('id', 0) > 0:
        district_col = cols['id'] - 1

    out = []
    for row in rows[header_i + 1:]:
        def val(k):
            ci = cols.get(k)
            return row[ci] if (ci is not None and ci < len(row)) else None
        idv = str(val('id') or '').strip()
        name = re.sub(r'[\s　]+', '', str(val('name') or '').strip())
        if not idv and not name:
            continue
        if re.match(r'^\s*\d+\.', idv):  # 2段目(日本語)ヘッダー行
            continue
        # 機器テスト用の見本行は除外
        if idv in SKIP_IDS or name.lower() in SKIP_NAMES:
            continue
        h, w = to_float(val('height')), to_float(val('weight'))
        if h is None or w is None:
            continue
        year, date = parse_year_date(val('testDate'))
        if not year:
            continue
        district = ''
        if district_col is not None and district_col < len(row) and row[district_col]:
            district = str(row[district_col]).strip()
        # 台帳の参加者ID形式(5桁数字)なら rosterId として直接突合に使う
        digits = re.sub(r'\D', '', idv)
        roster_id = digits.zfill(5) if idv and 4 <= len(digits) <= 5 and not idv.startswith('<') else None
        out.append({
            'ibId': idv, 'rosterId': roster_id, 'name': name, 'district': district,
            'sex': (str(val('gender') or '').strip().upper()[:1] or ''),
            'age': to_float(val('age')), 'height': h, 'weight': w,
            'year': year, 'testDate': date,
            'smm': to_float(val('smm')), 'smi': to_float(val('smi')),
            'fatPct': to_float(val('fatPct')), 'score': to_float(val('score')),
        })
    return out


# ---- Firestore --------------------------------------------------------------
def load_firestore(db):
    users = {}
    for d in db.collection('users').stream():
        u = d.to_dict()
        users[d.id] = {'sex': (u.get('sex') or 'F'),
                       'ward': (u.get('ward') or u.get('venueName') or ''),
                       'birth': u.get('birth'),
                       'name': re.sub(r'[\s　]+', '', str(u.get('name') or ''))}
    meas_by_year = {}
    for d in db.collection('measurements').stream():
        m = d.to_dict()
        uid, year = m.get('userId'), m.get('year')
        if uid is None or year is None:
            continue
        vals = m.get('values') or {}
        meas_by_year.setdefault(int(year), []).append({
            'ref': d.reference, 'userId': uid, 'date': m.get('date'),
            'height': to_float(vals.get('height')), 'weight': to_float(vals.get('weight')),
        })
    return users, meas_by_year


def build_profiles(users, meas_by_year):
    """利用者ごとの身体プロフィール。身長はほぼ不変なので、全年度の身長体重を突合材料にする。
       （令和4は台帳測定の多くに身長体重が無く、令和5は台帳に測定自体が無いため、
        「その年の測定」ではなく「その人」に対して突合する）"""
    prof = {uid: {**u, 'hw': {}} for uid, u in users.items()}
    for y, ms in meas_by_year.items():
        for m in ms:
            p = prof.get(m['userId'])
            if p is not None and (m['height'] is not None or m['weight'] is not None):
                p['hw'][y] = (m['height'], m['weight'])
    return prof


def resolve_user(ib, users, name_index, profiles):
    """InBody 1行 → 台帳の利用者を特定。確度の高い順: 台帳ID → 氏名 → 身体属性。"""
    if ib.get('rosterId') and ib['rosterId'] in users:
        return ib['rosterId'], 'id'
    if ib.get('name'):
        lst = name_index.get(ib['name'], [])
        if len(lst) == 1:
            return lst[0], 'name'
    # 身体属性: 性別・年齢(±1)・身長(全年度の記録と±2cm)・体重(最寄り年と比較)
    cands = []
    for uid, p in profiles.items():
        if ib['sex'] and p['sex'] and p['sex'] != ib['sex']:
            continue
        if ib['age'] and p['birth'] and abs((ib['year'] - p['birth']) - ib['age']) > 1:
            continue
        if not p['hw']:
            continue
        hds = [abs(h - ib['height']) for (h, _w) in p['hw'].values() if h is not None]
        if not hds or min(hds) > HEIGHT_TOL:
            continue
        wd, gap = None, None
        for y, (_h, w) in p['hw'].items():
            if w is None:
                continue
            g = abs(y - ib['year'])
            if gap is None or g < gap:
                gap, wd = g, abs(w - ib['weight'])
        if wd is not None and wd > (WEIGHT_TOL if gap == 0 else 4.0):
            continue  # 同年なら±1.5kg・別年なら±4kg まで
        score = min(hds) + (wd if wd is not None else 2.0)
        cands.append((uid, score, p['ward']))
    if not cands:
        return None, 'none'
    # 地区(行政区)が一致する候補を優先
    if ib['district']:
        byd = [c for c in cands if c[2] and c[2] == ib['district']]
        if byd:
            cands = byd
    if len(cands) == 1:
        return cands[0][0], 'person'
    cands.sort(key=lambda c: c[1])
    # 2位と十分差があれば最有力を採用、僅差なら保留（誤マッチ防止）
    if cands[0][1] + 0.8 <= cands[1][1]:
        return cands[0][0], 'person'
    return None, 'ambiguous'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--project', default='cruto-motion')
    ap.add_argument('--folder', default=DEFAULT_FOLDER_ID, help='InBody の Drive フォルダ ID')
    ap.add_argument('--dir', help='Drive を使わずローカルの xlsx フォルダから読む')
    ap.add_argument('--commit', action='store_true', help='実際に Firestore へ書き込む')
    args = ap.parse_args()

    import google.auth
    from google.cloud import firestore
    creds, _ = google.auth.default(scopes=SCOPES)

    if args.dir:
        files = sorted(glob.glob(os.path.join(args.dir, '*.xlsx')) + glob.glob(os.path.join(args.dir, '*.csv')))
        # ETL 出力の未突合CSVは取り込まない
        files = [f for f in files if not f.endswith('inbody-unmatched.csv')]
        print(f'ローカル {len(files)} ファイルを読み込みます')
    else:
        print(f'Drive フォルダ {args.folder} から取得中…')
        files = fetch_from_drive(args.folder, creds)
    if not files:
        print('InBody ファイルが見つかりませんでした')
        sys.exit(1)

    rows = []
    for f in files:
        try:
            r = parse_rows(read_rows(f))
        except Exception as e:  # noqa: BLE001
            print(f'  {os.path.basename(f)}: 読み込み失敗（{e}）— スキップ')
            continue
        print(f'  {os.path.basename(f)}: {len(r)} 行')
        rows += r
    # 同じ測定が xlsx/csv 重複で入るのを除く。同一人物でも年度が違えば別測定なので年を含める。
    seen, uniq = set(), []
    for ib in rows:
        key = (f"{ib['ibId']}|{ib['year']}" if ib['ibId']
               else f"{ib['year']}|{ib['district']}|{round(ib['height'])}|{round(ib['weight'])}|{ib['sex']}")
        if key in seen:
            continue
        seen.add(key)
        uniq.append(ib)
    rows = uniq
    print(f'InBody 合計 {len(rows)} 行（重複除去後）')

    db = firestore.Client(project=args.project, credentials=creds)
    users, meas_by_year = load_firestore(db)
    print(f'台帳: 利用者 {len(users)} 名 / 測定 {sum(len(v) for v in meas_by_year.values())} 件')
    # 氏名（空白除去）→ userId 一覧。同姓同名は一意にならないため一覧で持つ
    name_index = {}
    for uid, u in users.items():
        if u['name']:
            name_index.setdefault(u['name'], []).append(uid)
    profiles = build_profiles(users, meas_by_year)
    meas_by_key = {}
    for y, ms in meas_by_year.items():
        for m in ms:
            meas_by_key[(m['userId'], y)] = m

    # 年度別 availability（InBody行 / 台帳測定 / うち身長体重あり）— 突合率の当たりをつける
    from collections import Counter
    ib_year = Counter(ib['year'] for ib in rows)
    print('\n年度別  InBody行 / 台帳測定 / うち身長体重あり:')
    for y in sorted(set(list(ib_year.keys()) + list(meas_by_year.keys()))):
        mm = meas_by_year.get(y, [])
        hw = sum(1 for m in mm if m['height'] is not None and m['weight'] is not None)
        print(f'  {y}: {ib_year.get(y, 0)} / {len(mm)} / {hw}')

    # --- Pass 1: 各InBody行を利用者に解決 ---
    CONF = {'id': 3, 'name': 2, 'person': 1}
    method_count = {'id': 0, 'name': 0, 'person': 0, 'ambiguous': 0, 'none': 0}
    resolved, none_rows = [], []
    for ib in rows:
        uid, how = resolve_user(ib, users, name_index, profiles)
        method_count[how] += 1
        if uid:
            resolved.append((ib, uid, how))
        else:
            none_rows.append(ib)

    # --- Pass 2: (利用者, 年) ごとに最も確度の高い1行だけ採用 ---
    # ID/氏名で確定した行があればそれを優先（残りは同一人物の重複として破棄）。
    # 候補が身体属性マッチだけの競合は、誤紐付けを避けるため全て保留する。
    by_key = {}
    for ib, uid, how in resolved:
        by_key.setdefault((uid, ib['year']), []).append((ib, how))
    winners, dup_dropped, phys_conflict = [], 0, []
    for (uid, year), cl in by_key.items():
        cl.sort(key=lambda c: CONF[c[1]], reverse=True)
        if len(cl) == 1 or CONF[cl[0][1]] >= 2:
            winners.append((cl[0][0], uid, year))
            dup_dropped += len(cl) - 1
        else:
            phys_conflict += [ib for ib, _ in cl]
    unmatched = none_rows + phys_conflict

    # --- 書き込み ---
    n_attach = n_stub = n_fill = writes = 0
    batch, n_in_batch = db.batch(), 0

    def flush(force=False):
        nonlocal batch, n_in_batch
        if args.commit and (force or n_in_batch >= 400) and n_in_batch:
            batch.commit()
            batch, n_in_batch = db.batch(), 0

    for ib, uid, year in winners:
        inbody = {k: ib[k] for k in ('smm', 'smi', 'fatPct', 'score', 'height', 'weight', 'ibId', 'district', 'testDate')}
        inbody['source'] = 'lookinbody'
        m = meas_by_key.get((uid, year))
        if m:
            # 既存の測定に InBody を付与。台帳側に身長体重が無ければ InBody 実測で補完
            payload = {'inbody': inbody}
            fill = {}
            if m['height'] is None and ib['height'] is not None:
                fill['height'] = ib['height']
            if m['weight'] is None and ib['weight'] is not None:
                fill['weight'] = ib['weight']
            if fill:
                if ib['height'] and ib['weight']:
                    fill['bmi'] = round(ib['weight'] / (ib['height'] / 100) ** 2, 1)
                payload['values'] = fill
                n_fill += 1
            n_attach += 1
            if args.commit:
                batch.set(m['ref'], payload, merge=True)
                n_in_batch += 1; writes += 1; flush()
        else:
            # その年の測定が台帳に無い（例: 令和5年度シート欠落）→ InBody 単独の記録として登録。
            # 体力測定値をでっち上げないよう inbodyOnly を立て、アプリは InBody 欄にだけ表示する。
            vals = {'height': ib['height'], 'weight': ib['weight']}
            if ib['height'] and ib['weight']:
                vals['bmi'] = round(ib['weight'] / (ib['height'] / 100) ** 2, 1)
            n_stub += 1
            if args.commit:
                ref = db.collection('measurements').document(f'{uid}_{year}')
                batch.set(ref, {'userId': uid, 'year': year, 'date': ib['testDate'],
                                'inbodyOnly': True, 'values': vals, 'inbody': inbody,
                                'source': 'lookinbody'}, merge=True)
                n_in_batch += 1; writes += 1; flush()
    flush(force=True)

    print('\n=== 突合結果（本人特定の方法別）===')
    print(f'  台帳IDで確定  : {method_count["id"]}')
    print(f'  氏名で確定    : {method_count["name"]}')
    print(f'  身体属性で確定: {method_count["person"]}（性別+年齢+身長を全年度から照合+地区）')
    print(f'  曖昧(保留)    : {method_count["ambiguous"]}（候補が複数で確定できず）')
    print(f'  未一致        : {method_count["none"]}')
    print('  --- 取り込み内訳 ---')
    print(f'  同一人物の重複行を除外: {dup_dropped}')
    print(f'  身体属性の競合で保留  : {len(phys_conflict)}（どの人か確証が持てないため書き込まない）')
    print(f'  → 取り込む 利用者×年度: {len(winners)} 件')
    print(f'     ├ 既存の測定に付与       : {n_attach} 件（うち身長体重を補完 {n_fill} 件）')
    print(f'     └ InBody単独の記録を作成 : {n_stub} 件（台帳にその年の測定が無い分。例: 令和5年度）')
    print('  ' + (f'{writes} 件書き込みました' if args.commit else '（ドライラン・未書込）'))

    if unmatched:
        out = os.path.join(args.dir or '.', 'inbody-unmatched.csv')
        with open(out, 'w', newline='', encoding='utf-8-sig') as fp:
            w = csv.writer(fp)
            w.writerow(['ibId', 'rosterId', 'name', 'district', 'sex', 'age', 'height', 'weight', 'year', 'testDate'])
            for ib in unmatched:
                w.writerow([ib['ibId'], ib.get('rosterId') or '', ib.get('name') or '', ib['district'],
                            ib['sex'], ib['age'], ib['height'], ib['weight'], ib['year'], ib['testDate']])
        print(f'  未突合/曖昧 {len(unmatched)} 行 → {out}')
    if not args.commit:
        print('\n※ ドライランです。結果を確認し、問題なければ --commit を付けて再実行してください。')


if __name__ == '__main__':
    main()
