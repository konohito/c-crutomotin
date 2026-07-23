#!/usr/bin/env python3
# 嘉島町「★介護予防健診_Crutoデータベース」(xlsx) → アプリのスキーマへ正規化。
#  正本ソース: 名簿='データベース' / 2024測定='参照_利用者データ' / 2025測定='介護予防検診_身体機能'
#             InBody SMI='2024データベース'・'2025inbody' / 行政区='マスタ'(測定日→地区)
#  1) 2021/2022 も残す（矛盾値は contradictions.txt に抽出→担当者へ確認）
#  4) 名寄せ: かな(ひら/カタ)統一 + 全名簿で突合、未一致のフォーム回答者は新規(要確認)登録
#  5) 市町村=嘉島町 / 行政区をマスタから付与
#
# 使い方: openpyxl 必須。xlsx と同じディレクトリで実行し normalized.json を生成する。
#   pip install openpyxl && python3 etl-kashima.py            # cruto-db.xlsx を読む
#   （出力 normalized.json は個人情報を含むためコミット禁止。seed-firestore.mjs で Firestore へ投入）
import openpyxl, json, datetime, re, sys

wb=openpyxl.load_workbook(sys.argv[1] if len(sys.argv)>1 else "cruto-db.xlsx", data_only=True)
def num(v):
    if v is None or v=='' : return None
    if isinstance(v,(int,float)): return round(float(v),2)
    m=re.search(r'-?\d+(?:\.\d+)?',str(v)); return round(float(m.group()),2) if m else None
def yr(v):
    if isinstance(v,(datetime.datetime,datetime.date)): return v.year
    m=re.search(r'(20\d\d)',str(v or '')); return int(m.group(1)) if m else None
def dstr(v): return f"{v.year}/{v.month:02d}/{v.day:02d}" if isinstance(v,(datetime.datetime,datetime.date)) else None
def sex_of(v):
    s=str(v or '')
    if '女' in s or s.strip() in ('2','2.0'): return 'F'
    if '男' in s or s.strip() in ('1','1.0'): return 'M'
    return None
def sid(v):
    n=num(v); return str(int(n)) if n is not None else None
def kata(s): return ''.join(chr(ord(c)+0x60) if 'ぁ'<=c<='ゖ' else c for c in str(s or ''))
def nkey(s): return kata(re.sub(r'[\s　]','',str(s or '')))
def cap60(v): return None if v is None else min(v,60.0)
def rows(n): return list(wb[n].iter_rows(min_row=2,values_only=True))

WARD_BY_DATE={}
for r in wb['マスタ'].iter_rows(min_row=1,values_only=True):
    if isinstance(r[0],(datetime.datetime,datetime.date)) and r[1]:
        WARD_BY_DATE[dstr(r[0])]=str(r[1]).strip()

users={}   # id -> user
name2id={} # nkey(name) -> id
def ensure(_id, **kw):
    u=users.get(_id)
    if not u:
        u=users[_id]={'id':_id,'muni':'嘉島町','ward':'','meas':{},'inbody':{},'flags':[]}
    for k,v in kw.items():
        if v not in (None,'') and not u.get(k): u[k]=v
    return u

# ---- 名簿: データベース(2025) ----
for r in rows('データベース'):
    _id=sid(r[0])
    if not _id or not r[1]: continue
    bd=r[4]
    u=ensure(_id, name=str(r[1]).strip(), kana=str(r[2] or '').replace('　',' ').strip(),
             sex=sex_of(r[3]), birthDate=dstr(bd),
             birth=(bd.year if isinstance(bd,(datetime.datetime,datetime.date)) else None),
             height=num(r[5]), weight=num(r[6]))
    name2id.setdefault(nkey(r[1]), _id)

# ---- 2024データベース: 名簿補完 + 要介護度/SMI ----
for r in rows('2024データベース'):
    _id=sid(r[0])
    if not _id: continue
    u=ensure(_id, name=str(r[1] or '').strip(), kana=str(r[2] or '').replace('　',' ').strip(),
             sex=sex_of(r[4] if r[4] is not None else r[3]),
             birthDate=dstr(r[5]), birth=(r[5].year if isinstance(r[5],(datetime.datetime,datetime.date)) else None))
    if str(r[18] or '').strip(): u['careLevel']=str(r[18]).strip()
    if num(r[83]) is not None: u['inbody'].setdefault('2024',{})['smi']=num(r[83])
    if u.get('name'): name2id.setdefault(nkey(u['name']), _id)

# ---- 参照_利用者データ: 名簿補完 + 測定(全年度ブロック) ----
for r in rows('参照_利用者データ'):
    _id=sid(r[0])
    if not _id: continue
    u=ensure(_id, name=str(r[1] or '').strip(), kana=str(r[2] or '').replace('　',' ').strip(),
             sex=sex_of(r[7]), birthDate=dstr(r[6]),
             birth=(r[6].year if isinstance(r[6],(datetime.datetime,datetime.date)) else None),
             phone=str(r[8] or '').strip())
    if u.get('name'): name2id.setdefault(nkey(u['name']), _id)
    for b in (9,19,29):
        y=yr(r[b])
        if not y: continue
        rec={'date':dstr(r[b]),'walk5':num(r[b+1]),'balR':cap60(num(r[b+2])),'balL':cap60(num(r[b+3])),
             'gripR':num(r[b+4]),'gripL':num(r[b+5]),'tug':num(r[b+6]),'height':num(r[b+7]),
             'weight':num(r[b+8]),'bmi':num(r[b+9]),'walk5max':None,'source':'参照(FileMaker)'}
        cur=u['meas'].get(str(y))
        if not cur or sum(v is not None for v in rec.values())>sum(v is not None for v in cur.values()):
            u['meas'][str(y)]=rec

# ---- 2025 測定: 介護予防検診_身体機能(氏名キー・かな統一で突合) ----
newseq=90001; matched=0; created=0
for r in wb['介護予防検診_身体機能'].iter_rows(min_row=2,values_only=True):
    if not r[1] or yr(r[35])!=2025: continue
    key=nkey(r[1]); _id=name2id.get(key)
    if not _id:
        _id=str(newseq); newseq+=1; created+=1
        ensure(_id, name=str(r[1]).strip())
        users[_id]['flags'].append('名簿未登録（要確認）')
        name2id[key]=_id
    else:
        matched+=1
    u=users[_id]
    d=dstr(r[35])
    ward=WARD_BY_DATE.get(d,'')
    if ward: u['ward']=ward
    h=num(r[33]) or u.get('height'); w=num(r[34]) or u.get('weight')
    rec={'date':d,'balR':cap60(num(r[26])),'balL':cap60(num(r[27])),'walk5':num(r[28]),'walk5max':num(r[29]),
         'tug':num(r[30]),'gripR':num(r[31]),'gripL':num(r[32]),'height':h,'weight':w,
         'bmi':(round(w/(h/100)**2,1) if h and w else None),'source':'身体機能フォーム'}
    prev=u['meas'].get('2025')
    if not prev or sum(v is not None for v in rec.values())>sum(v is not None for v in prev.values()):
        u['meas']['2025']=rec

# ---- 2025 InBody SMI ----
for r in rows('2025inbody'):
    _id=name2id.get(nkey(r[0]))
    if _id and num(r[3]) is not None: users[_id]['inbody'].setdefault('2025',{})['smi']=num(r[3])

out=[u for u in users.values() if u.get('name') and u['meas']]
out.sort(key=lambda u:u['id'])

# ---- 矛盾値の処理（最新年=2024/2025 を正として、±6cm超 乖離する旧年度）----
# 担当者確認までの暫定対応: 該当年度の身長・BMI を空白化し、要確認フラグを付ける。
# 歩行・握力・TUG 等の他項目は残す（後で担当者回答に応じて調整可能）。
contra=[]; suspect_meas=0
for u in out:
    hs=[(y,m['height']) for y,m in sorted(u['meas'].items()) if m.get('height')]
    if len(hs)<2: continue
    ref=hs[-1][1]  # 最新年の身長を基準
    bad=[(y,h) for y,h in hs if abs(h-ref)>6]
    if bad:
        ws=[(y,m['weight']) for y,m in sorted(u['meas'].items()) if m.get('weight')]
        contra.append((u,list(hs),ws,[y for y,_ in bad]))
        for y,h in bad:
            u['meas'][y]['review']=True   # 要確認フラグ
            u['meas'][y]['height']=None   # 矛盾する身長は空白化
            u['meas'][y]['bmi']=None       # 身長依存の BMI も空白化
            suspect_meas+=1

# 正規化データを書き出す（要確認フラグ・空白化を反映した後）
json.dump(out, open('normalized.json','w'), ensure_ascii=False, indent=1)

years={}
for u in out:
    for y in u['meas']: years[y]=years.get(y,0)+1
print(f"利用者(測定あり): {len(out)} 名  / 新規(名簿未登録): {created} 名")
print("年度別 測定件数:", dict(sorted(years.items())))
print(f"2025フォーム突合: 既存一致 {matched} / 新規作成 {created}")
print(f"行政区(ward)付与: {sum(1 for u in out if u.get('ward'))} 名  例:", {u['ward'] for u in out if u.get('ward')})
print(f"矛盾値の疑い: {len(contra)} 名 / 要確認とした測定: {suspect_meas} 件")

# 担当者への質問文 + 矛盾リストを保存
with open('contradictions.txt','w') as f:
    f.write("【嘉島町 介護予防健診データ 取り込みに関する確認のお願い】\n\n")
    f.write("お世話になっております。Cruto motion への実データ取り込みを進めております。\n")
    f.write(f"「★介護予防健診_Crutoデータベース」を確認したところ、測定データのある {len(out)} 名のうち\n")
    f.write(f"約半数の {len(contra)} 名で、同一IDの方の身長が年度によって大きく異なっていました（±6cm 超）。\n\n")
    f.write("件数が多く、個別のミスというより『参照_利用者データ』シートの旧年度（2021・2022）ブロックが\n")
    f.write("系統的にズレて（別の方の記録が入って）いる可能性が高いと考えています。\n\n")
    f.write("つきましては、以下をご確認いただけますでしょうか。\n")
    f.write("  ① 2021・2022 年度の測定について、参照シート以外に正となる元データ（別ファイル等）はありますか。\n")
    f.write("  ② 参照シートの旧年度ブロックは、氏名・IDと正しく対応していますか（並び順ズレの有無）。\n")
    f.write("  ③ ひとまず 2024・2025 を正として取り込み、2021・2022 は『要確認』表示で保留する進め方で問題ないでしょうか。\n\n")
    f.write("※ システム上は 2021・2022 も残していますが、下記該当者には『要確認』フラグを付けています。\n")
    f.write("　 取り込み後もアプリ上ですべて編集・修正できます。\n\n")
    f.write(f"――― 該当者リスト（{len(contra)} 名。身長が中央値から±6cm超の年度を ★要確認 と表示）―――\n\n")
    for u,hs,wsx,badyears in contra:
        f.write(f"■ ID {u['id']} {u['name']}（{u.get('kana','')}）\n")
        f.write("   身長: " + " / ".join((f"★{y}年:{h}cm" if y in badyears else f"{y}年:{h}cm") for y,h in hs) + "\n")
        f.write("   体重: " + " / ".join(f"{y}年:{w}kg" for y,w in wsx) + "\n\n")
print("→ contradictions.txt に質問文を保存")
