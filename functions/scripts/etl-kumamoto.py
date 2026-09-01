#!/usr/bin/env python3
# 熊本市「個人管理台帳」書き出し (xlsx) → アプリのスキーマへ正規化。
#   1 行 = 1 名 × 1 年度。開始時・終了時の 2 回の測定と、基本チェックリスト(0/1)を持つ。
#
# 年度の対応(重要):
#   アプリの「今年度」は D.CUR = 2025(画面・取り込みはすべて 2025 に記録される)。
#   熊本市の台帳の年度は 4 月区切りで、2026年度 = いま進行中のシーズン。
#   よって アプリ年度 = 測定日の年度(4月区切り) - 1 で取り込む。
#     例: 2025-09-03 の測定(台帳では 2025 年度の開始時) → アプリの 2024 = 「昨年」
#         2026-04 以降の測定(2026 年度) → アプリの 2025 = 「今年度」
#
# ID の対応:
#   台帳の ID は 4 桁(1001〜)。嘉島町の既存 ID と衝突しないよう「2」を前置した
#   5 桁(21001〜)をアプリの ID にする。元の ID は extId に残す。
#
# 使い方: openpyxl 必須。
#   pip install openpyxl && python3 etl-kumamoto.py 台帳.xlsx
#   → normalized-kumamoto.json を生成(個人情報を含むためコミット禁止)。
#   → node scripts/seed-firestore.mjs normalized-kumamoto.json で Firestore へ投入。
import openpyxl, json, datetime, re, sys

wb = openpyxl.load_workbook(sys.argv[1] if len(sys.argv) > 1 else "kumamoto.xlsx", data_only=True)
ws = wb[wb.sheetnames[0]]

def num(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)): return round(float(v), 2)
    m = re.search(r'-?\d+(?:\.\d+)?', str(v))
    return round(float(m.group()), 2) if m else None
def dstr(v):
    return f"{v.year}/{v.month:02d}/{v.day:02d}" if isinstance(v, (datetime.datetime, datetime.date)) else None
def sex_of(v):
    s = str(v or '')
    return 'F' if '女' in s else 'M' if '男' in s else None
def cap60(v):
    return None if v is None else min(v, 60.0)
def app_year(v):
    # 測定日 → アプリの年度(= 4 月区切りの年度 - 1)
    if not isinstance(v, (datetime.datetime, datetime.date)): return None
    fy = v.year - (1 if v.month < 4 else 0)
    return fy - 1

# 基本チェックリストの点数化方向(kihon.js KCL_QUESTIONS と一致させること)。
# 台帳は 0/1(1=該当=1点)。アプリは raw {'設問No': 'yes'|'no'} で持つ。No.12(BMI)は派生のため除外。
POINT_ON = {1:'no',2:'no',3:'no',4:'no',5:'no',6:'no',7:'no',8:'no',9:'yes',10:'yes',
            11:'yes',13:'yes',14:'yes',15:'yes',16:'no',17:'yes',18:'yes',19:'no',20:'yes',
            21:'yes',22:'yes',23:'yes',24:'yes',25:'yes'}

users = {}
def ensure(_id, **kw):
    u = users.get(_id)
    if not u:
        u = users[_id] = {'id': _id, 'meas': {}, 'kcl': {}, 'flags': []}
    for k, v in kw.items():
        if v not in (None, '') and not u.get(k): u[k] = v
    return u

def merge_meas(u, y, rec):
    """同じアプリ年度に 2 つの測定が重なったら(C型の開始と終了など)、
       新しい日付を正として、欠けている値だけ古い方から補う。"""
    y = str(y)
    cur = u['meas'].get(y)
    if not cur:
        u['meas'][y] = rec; return
    older, newer = (cur, rec) if (rec.get('date') or '') >= (cur.get('date') or '') else (rec, cur)
    merged = dict(older)
    for k, v in newer.items():
        if v is not None: merged[k] = v
    u['meas'][y] = merged

skipped = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    rid = num(r[1])
    if rid is None or not r[2]:
        skipped += 1; continue
    _id = '2' + str(int(rid)).zfill(4)          # 1001 → 21001(嘉島町の ID と重ならない 5 桁)
    kuName = str(r[10] or '').strip()           # 東区 / 西区 / 南区
    u = ensure(_id,
               extId=str(int(rid)),
               name=str(r[2]).strip(), kana=str(r[3] or '').replace('　', ' ').strip(),
               sex=sex_of(r[4]), birthDate=dstr(r[5]),
               birth=(r[5].year if isinstance(r[5], (datetime.datetime, datetime.date)) else None),
               muni=('熊本市' + kuName) if kuName else '熊本市',
               muniName=('熊本市' + kuName) if kuName else '熊本市',
               region='熊本市圏域',
               ward=str(r[11] or '').replace('　', ' ').strip(),
               careLevel=(str(r[15]).strip() if r[15] not in (None, '', 'なし') else ''))

    h, w, bmi = num(r[12]), num(r[13]), num(r[14])

    # 開始時の測定(+ 身長体重は開始時のもの・BMI も)
    ys = app_year(r[52])
    if ys is not None:
        merge_meas(u, ys, {
            'date': dstr(r[52]),
            'balR': cap60(num(r[61])), 'balL': None,
            'tug': num(r[63]), 'walk5': num(r[65]), 'walk5max': num(r[67]),
            'gripR': num(r[69]), 'gripL': None,
            'height': h, 'weight': w, 'bmi': bmi,
            'source': '個人管理台帳(熊本市)・開始時',
        })
        # 基本チェックリストは開始時に実施 → 開始時の年度に付ける
        raw = {}
        for no, po in POINT_ON.items():
            v = r[19 + no - 1]
            if v in (0, 1):
                raw[str(no)] = po if v == 1 else ('no' if po == 'yes' else 'yes')
        if raw:
            u['kcl'][str(ys)] = raw
    # 終了時の測定(身長は変わらないので引き継ぐ。体重は測っていないため入れない)
    ye = app_year(r[53])
    if ye is not None:
        merge_meas(u, ye, {
            'date': dstr(r[53]),
            'balR': cap60(num(r[62])), 'balL': None,
            'tug': num(r[64]), 'walk5': num(r[66]), 'walk5max': num(r[68]),
            'gripR': num(r[70]), 'gripL': None,
            'height': h, 'weight': None, 'bmi': None,
            'source': '個人管理台帳(熊本市)・終了時',
        })

out = [u for u in users.values() if u.get('name')]
out.sort(key=lambda u: u['id'])
json.dump(out, open('normalized-kumamoto.json', 'w'), ensure_ascii=False, indent=1)

years = {}
for u in out:
    for y in u['meas']: years[y] = years.get(y, 0) + 1
kclN = sum(1 for u in out for _ in u['kcl'])
names = {}
for u in out: names[u['name']] = names.get(u['name'], 0) + 1
dups = {n: c for n, c in names.items() if c > 1}
print(f"利用者: {len(out)} 名(読み飛ばし {skipped} 行)")
print("アプリ年度別の測定件数:", dict(sorted(years.items())), "(2024=昨年 / 2025=今年度)")
print(f"基本チェックリスト: {kclN} 年度ぶん")
print("行政区:", sorted({u.get('ward') for u in out if u.get('ward')}))
if dups: print("⚠ 同姓同名(別 ID。取り込み後に台帳でご確認を):", dups)
